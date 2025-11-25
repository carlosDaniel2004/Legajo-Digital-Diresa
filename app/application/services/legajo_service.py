# app/application/services/legajo_service.py
# Importa la librería para calcular hashes de archivos.
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import io
from flask import current_app
from datetime import datetime, timedelta
import secrets
import string
import logging

logger = logging.getLogger(__name__)


# Define el servicio que contiene la lógica de negocio para los legajos.
class LegajoService:
    # El constructor inyecta las dependencias del repositorio de personal y el servicio de auditoría.
    def __init__(self, personal_repository, audit_service, usuario_service=None):
        self._personal_repo = personal_repository
        self._audit_service = audit_service
        self._usuario_service = usuario_service

    # --- MÉTODOS DE CONSULTA (GETTERS) ---

    def get_tipos_documento_by_seccion(self, seccion_id):
        """Orquesta la obtención de tipos de documento filtrados por sección."""
        return self._personal_repo.get_tipos_documento_by_seccion(seccion_id)

    def get_document_for_download(self, document_id):
        """Recupera un documento de la base de datos y lo prepara para la descarga."""
        document_row = self._personal_repo.find_document_by_id(document_id)
        if not document_row:
            return None
        # El SP devuelve una fila con (nombre_archivo, archivo_binario)
        return {"filename": document_row[0], "data": document_row[1]}

    def check_if_dni_exists(self, dni):
        """Orquesta la verificación de la existencia de un DNI."""
        return self._personal_repo.check_dni_exists(dni)

    def get_all_personal_paginated(self, page, per_page, filters=None):
        """Obtiene una lista paginada y filtrada de personal."""
        return self._personal_repo.get_all_paginated(page, per_page, filters)

    def get_personal_details(self, personal_id, current_user):
        """
        Obtiene todos los detalles del legajo de una persona por su ID,
        aplicando reglas de control de acceso.
        """
        legajo = self._personal_repo.get_full_legajo_by_id(personal_id)
        if not legajo:
            return None

        # Seguridad: Los roles 'RRHH', 'Sistemas' y 'AdministradorLegajos' tienen permitido ver cualquier legajo.
        # Se mantiene la estructura por si se necesita añadir lógica de permisos más granular en el futuro.
        if current_user.rol not in ['RRHH', 'Sistemas', 'AdministradorLegajos']:
            # Si el rol no es uno de los permitidos, se deniega el acceso.
            # (Actualmente, los decoradores de ruta ya previenen esto, pero es una doble capa de seguridad).
            raise PermissionError("No tiene permiso para ver este legajo.")

        return legajo

    def get_documents_by_personal_id(self, personal_id):
        """Obtiene los documentos de un empleado."""
        return self._personal_repo.find_documents_by_personal_id(personal_id)

    # --- MÉTODOS PARA POBLAR FORMULARIOS ---

    def get_unidades_for_select(self):
        return self._personal_repo.get_unidades_for_select()

    def get_secciones_for_select(self):
        return self._personal_repo.get_secciones_for_select()

    def get_tipos_documento_for_select(self):
        return self._personal_repo.get_tipos_documento_for_select()

    # --- MÉTODOS DE OPERACIONES (CUD) ---

    def register_new_personal(self, form_data, creating_user_id):
        """
        Registra un nuevo empleado y audita la acción.
        Genera automáticamente usuario con rol Personal.
        Username y contraseña son el DNI del empleado.
        """
        try:
            # 1. Crear el personal
            new_personal_id = self._personal_repo.create(form_data)
            
            # 2. Usar DNI como username y contraseña
            dni = form_data.get('dni', '').strip()
            if not dni:
                raise ValueError("El DNI es requerido para crear el usuario")
            
            username = dni
            password = dni
            email = form_data.get('email', '')
            
            # 3. Obtener ID del rol "Personal"
            id_rol_personal = self._get_personal_role_id()
            
            if not id_rol_personal:
                logger.error("No se encontró ningún rol disponible en la base de datos")
                raise Exception("Error de configuración: No hay roles disponibles en el sistema")
            
            # 4. Crear usuario con los datos generados
            if self._usuario_service:
                user_data = {
                    'username': username,
                    'email': email,
                    'password': password,
                    'id_rol': id_rol_personal,
                    'id_personal': new_personal_id  # Asociar con el personal creado
                }
                # create_user devuelve (mensaje, tipo) -> verificar resultado
                result = self._usuario_service.create_user(user_data)
                # Si la función devuelve tupla (mensaje, tipo)
                if isinstance(result, tuple) and len(result) >= 2:
                    mensaje, tipo = result[0], result[1]
                else:
                    mensaje, tipo = str(result), 'unknown'

                if tipo != 'success':
                    logger.error(f"Error creando usuario automáticamente: {mensaje}")
                    # Informar y revertir: lanzar excepción para que el caller lo maneje
                    raise Exception(f"No se pudo crear el usuario automáticamente: {mensaje}")

                logger.info(f"Usuario '{username}' creado automáticamente para personal ID {new_personal_id} con DNI")
            else:
                logger.warning("Usuario service no disponible - no se creó el usuario")
            
            # 5. Auditar la acción
            audit_data = form_data.copy()
            audit_data['username_generado'] = username
            audit_data['nota'] = f"Usuario creado automáticamente con DNI {dni} como username y contraseña"
            self._audit_service.log(
                creating_user_id, 
                'Personal', 
                'CREAR', 
                f"Se creó el legajo para el DNI {form_data['dni']} con usuario {username}", 
                audit_data
            )
            
            return new_personal_id
            
        except Exception as e:
            logger.error(f"Error en register_new_personal: {str(e)}")
            raise

    def _generate_username(self, nombres: str, apellidos: str) -> str:
        """
        Genera un username único a partir del nombre y apellido.
        Formato: primeraletra_apellido (ej: c_hernandez)
        """
        try:
            # Obtener primer nombre y primer apellido, sin espacios
            primer_nombre = nombres.strip().split()[0].lower() if nombres else "user"
            primer_apellido = apellidos.strip().split()[0].lower() if apellidos else "personal"
            
            base_username = f"{primer_nombre[0]}_{primer_apellido}"
            username = base_username
            
            # Si el username ya existe, agregar números
            counter = 1
            while self._usuario_service and self._usuario_service._usuario_repo.find_by_username(username):
                username = f"{base_username}{counter}"
                counter += 1
            
            return username
            
        except Exception as e:
            logger.error(f"Error generando username: {e}")
            # Fallback: generar username aleatorio
            return f"user_{secrets.token_hex(4)}"

    def _generate_password(self, length: int = 12) -> str:
        """
        Genera una contraseña segura aleatoria.
        Incluye mayúsculas, minúsculas, números y símbolos.
        """
        uppercase = string.ascii_uppercase
        lowercase = string.ascii_lowercase
        digits = string.digits
        symbols = "!@#$%^&*-_=+"
        
        # Asegurar que haya al menos uno de cada tipo
        password_chars = [
            secrets.choice(uppercase),
            secrets.choice(lowercase),
            secrets.choice(digits),
            secrets.choice(symbols)
        ]
        
        # Llenar el resto de caracteres aleatoriamente
        all_chars = uppercase + lowercase + digits + symbols
        password_chars += [secrets.choice(all_chars) for _ in range(length - 4)]
        
        # Mezclar la contraseña
        random_password = ''.join(secrets.SystemRandom().sample(password_chars, len(password_chars)))
        
        return random_password

    def _get_personal_role_id(self) -> int:
        """
        Obtiene el ID del rol 'Personal' de la BD.
        Si no existe, busca un rol alternativo válido.
        """
        try:
            from app.database.connector import get_db_read
            conn = get_db_read()
            cursor = conn.cursor()
            
            # Primero intentar obtener el rol 'Personal'
            cursor.execute("SELECT id_rol FROM roles WHERE nombre_rol = 'Personal'")
            row = cursor.fetchone()
            
            if row:
                logger.info(f"Rol 'Personal' encontrado con ID: {row[0]}")
                return row[0]
            
            # Si no existe, usar el rol alternativo 'RRHH' (válido para empleados)
            logger.warning("Rol 'Personal' no encontrado en la BD, usando 'RRHH' como alternativa")
            cursor.execute("SELECT id_rol FROM roles WHERE nombre_rol = 'RRHH'")
            row = cursor.fetchone()
            
            if row:
                logger.info(f"Usando rol alternativo 'RRHH' con ID: {row[0]}")
                return row[0]
            
            # Si tampoco existe RRHH, obtener cualquier rol disponible
            logger.warning("Rol 'RRHH' tampoco encontrado, obteniendo primer rol disponible")
            cursor.execute("SELECT TOP 1 id_rol FROM roles ORDER BY id_rol")
            row = cursor.fetchone()
            
            if row:
                logger.info(f"Usando primer rol disponible con ID: {row[0]}")
                return row[0]
            
            logger.error("No hay ningún rol disponible en la base de datos")
            return None
            
        except Exception as e:
            logger.error(f"Error obteniendo ID del rol: {e}")
            return None

    def upload_document_to_personal(self, form_data, file_storage, current_user_id):
        """Gestiona la validación y subida de un nuevo documento."""
        if not file_storage or not file_storage.filename:
            raise ValueError("No se proporcionó ningún archivo para subir.")

        filename = file_storage.filename
        
        allowed_extensions = current_app.config['ALLOWED_EXTENSIONS']
        if '.' not in filename or filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
            raise ValueError(f"Tipo de archivo no permitido. Solo se aceptan: {', '.join(allowed_extensions)}")

        file_bytes = file_storage.read()
        if len(file_bytes) > current_app.config['MAX_CONTENT_LENGTH']:
            max_size_mb = current_app.config['MAX_CONTENT_LENGTH'] / (1024 * 1024)
            raise ValueError(f"El archivo es demasiado grande. El tamaño máximo es de {max_size_mb:.0f} MB.")
        
        file_storage.seek(0)
        
        file_hash = hashlib.sha256(file_bytes).hexdigest()
        
        doc_data = form_data.copy()
        doc_data['nombre_archivo'] = filename
        doc_data['hash_archivo'] = file_hash
        id_personal = doc_data.get('id_personal')

        self._personal_repo.add_document(doc_data, file_bytes)
        
        self._audit_service.log(
            current_user_id,
            'Documentos',
            'SUBIR',
            f"Subió el archivo '{filename}' al legajo del personal ID {id_personal}"
        )

    def delete_personal_by_id(self, personal_id, deleting_user_id):
        """Desactiva un legajo de personal, su usuario asociado (si existe), y audita la acción."""
        persona = self._personal_repo.find_by_id(personal_id)
        if not persona:
            raise ValueError("La persona que intenta eliminar no existe.")

        self._personal_repo.delete_by_id(personal_id)
        
        # Si existe un usuario asociado a este personal, desactivarlo también
        if self._usuario_service:
            try:
                # Buscar usuario por DNI o email
                usuario = self._usuario_service._usuario_repo.find_by_email(persona.email) if hasattr(persona, 'email') and persona.email else None
                if usuario:
                    self._usuario_service._usuario_repo.deactivate_user(usuario.id)
                    self._audit_service.log(
                        deleting_user_id,
                        'Usuario',
                        'DESACTIVAR (Cascada)',
                        f"Usuario asociado a personal DNI {persona.dni} fue desactivado automáticamente"
                    )
            except Exception as e:
                # Log del error pero no detiene la desactivación del personal
                self._audit_service.log(
                    deleting_user_id,
                    'Usuario',
                    'ERROR_DESACTIVACION',
                    f"Error al desactivar usuario de personal DNI {persona.dni}: {str(e)}"
                )
        
        self._audit_service.log(
            deleting_user_id,
            'Personal',
            'ELIMINAR (Desactivar)',
            f"Se desactivó el legajo del personal con DNI {persona.dni}"
        )

    def activate_personal_by_id(self, personal_id, activating_user_id):
        """Reactiva un legajo de personal, su usuario asociado (si existe), y audita la acción."""
        persona = self._personal_repo.find_by_id(personal_id)
        if not persona:
            raise ValueError("La persona que intenta activar no existe.")

        self._personal_repo.activate_by_id(personal_id)
        
        # Si existe un usuario asociado a este personal, reactivarlo también
        if self._usuario_service:
            try:
                # Buscar usuario por DNI o email
                usuario = self._usuario_service._usuario_repo.find_by_email(persona.email) if hasattr(persona, 'email') and persona.email else None
                if usuario:
                    self._usuario_service._usuario_repo.activate_user(usuario.id)
                    self._audit_service.log(
                        activating_user_id,
                        'Usuario',
                        'ACTIVAR (Cascada)',
                        f"Usuario asociado a personal DNI {persona.dni} fue reactivado automáticamente"
                    )
            except Exception as e:
                # Log del error pero no detiene la reactivación del personal
                self._audit_service.log(
                    activating_user_id,
                    'Usuario',
                    'ERROR_REACTIVACION',
                    f"Error al reactivar usuario de personal DNI {persona.dni}: {str(e)}"
                )
        
        self._audit_service.log(
            activating_user_id,
            'Personal',
            'ACTIVAR (Reactivar)',
            f"Se reactivó el legajo del personal con DNI {persona.dni}"
        )
    def delete_document_by_id(self, document_id, deleting_user_id):
        """Orquesta la eliminación lógica de un documento y lo audita."""
        self._personal_repo.delete_document_by_id(document_id)
        self._audit_service.log(
            deleting_user_id,
            'Documentos',
            'ELIMINAR (Lógico)',
            f"Se marcó como eliminado el documento con ID {document_id}"
        )

    def process_bulk_upload(self, file_storage, creating_user_id):
        """
        Procesa un archivo Excel para la carga masiva de personal.
        Valida cada fila y registra a los nuevos empleados.
        """
        import openpyxl
        
        workbook = openpyxl.load_workbook(file_storage)
        sheet = workbook.active
        
        # Se asume que la primera fila es el encabezado.
        headers = [cell.value for cell in sheet[1]]
        
        # Columnas esperadas en la plantilla.
        expected_headers = [
            "DNI", "Nombres", "Apellidos", "Sexo", "FechaNacimiento", "Telefono",
            "Email", "Direccion", "EstadoCivil", "Nacionalidad", "UnidadAdministrativa",
            "FechaIngreso"
        ]

        # Validación simple de encabezados.
        if headers[:len(expected_headers)] != expected_headers:
            raise ValueError("El formato del archivo Excel es incorrecto. Las columnas no coinciden con la plantilla.")

        unidades_map = {nombre: id_ for id_, nombre in self._personal_repo.get_unidades_for_select()}

        registros_exitosos = 0
        registros_fallidos = 0
        errores = []

        # Itera sobre las filas, omitiendo el encabezado.
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))
            
            try:
                # --- Validación de Datos ---
                if not all([row_data.get('DNI'), row_data.get('Nombres'), row_data.get('Apellidos')]):
                    raise ValueError("DNI, Nombres y Apellidos son obligatorios.")
                
                unidad_nombre = row_data.get('UnidadAdministrativa')
                if not unidad_nombre or unidad_nombre not in unidades_map:
                    raise ValueError(f"La unidad administrativa '{unidad_nombre}' no es válida.")

                # --- Mapeo de Datos para el Formulario ---
                form_data = {
                    'dni': str(row_data['DNI']),
                    'nombres': row_data['Nombres'],
                    'apellidos': row_data['Apellidos'],
                    'sexo': row_data.get('Sexo'),
                    'fecha_nacimiento': row_data.get('FechaNacimiento'),
                    'telefono': row_data.get('Telefono'),
                    'email': row_data.get('Email'),
                    'direccion': row_data.get('Direccion'),
                    'estado_civil': row_data.get('EstadoCivil'),
                    'nacionalidad': row_data.get('Nacionalidad', 'Peruana'),
                    'id_unidad': unidades_map[unidad_nombre],
                    'fecha_ingreso': row_data.get('FechaIngreso')
                }
                
                # Llama al método de registro existente.
                self.register_new_personal(form_data, creating_user_id)
                registros_exitosos += 1

            except Exception as e:
                registros_fallidos += 1
                errores.append(f"Fila {row_index}: {e}")
        
        return {"exitosos": registros_exitosos, "fallidos": registros_fallidos, "errores": errores}

    def generate_bulk_upload_template(self, unidades):
        """
        Genera una plantilla de Excel con las columnas necesarias y validaciones de datos.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Plantilla de Carga de Personal"

        headers = [
            "DNI", "Nombres", "Apellidos", "Sexo", "FechaNacimiento", "Telefono",
            "Email", "Direccion", "EstadoCivil", "Nacionalidad", "UnidadAdministrativa",
            "FechaIngreso"
        ]
        ws.append(headers)

        # Estilo para el encabezado.
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # --- Validación de Datos en Excel ---
        # Validación para la columna de Sexo (D).
        dv_sexo = DataValidation(type="list", formula1='"M,F"', allow_blank=True)
        dv_sexo.error = "Por favor, ingrese 'M' para Masculino o 'F' para Femenino."
        dv_sexo.errorTitle = "Valor no válido"
        ws.add_data_validation(dv_sexo)
        dv_sexo.add('D2:D1000')

        # Validación para la columna de Unidad Administrativa (K).
        nombres_unidades = [nombre for _, nombre in unidades]
        formula_unidades = f'"{",".join(nombres_unidades)}"'
        dv_unidad = DataValidation(type="list", formula1=formula_unidades, allow_blank=False)
        dv_unidad.error = "Por favor, seleccione una unidad de la lista."
        dv_unidad.errorTitle = "Unidad no válida"
        ws.add_data_validation(dv_unidad)
        dv_unidad.add('K2:K1000')

        # Ajustar ancho de columnas.
        for i, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = len(header) + 5

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return excel_stream

    # --- MÉTODOS DE REPORTES Y ESTADO ---
    
    def generate_general_report_excel(self):
        """Genera un reporte general de personal en un archivo Excel."""
        personal_data = self._personal_repo.get_all_for_report()
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte General de Personal"

        headers = [
            "DNI", "Apellidos", "Nombres", "Sexo", "Fecha de Nacimiento", "Email",
            "Teléfono", "Unidad Administrativa", "Fecha de Ingreso", "Estado",
            "Último Cargo", "Último Tipo de Contrato", "Modalidad", "Sueldo", "Resolución"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for persona in personal_data:
            row_data = [
                persona.get('dni'), persona.get('apellidos'), persona.get('nombres'),
                persona.get('sexo'), persona.get('fecha_nacimiento'), persona.get('email'),
                persona.get('telefono'), persona.get('nombre_unidad'),
                persona.get('fecha_ingreso'), 'Activo' if persona.get('activo') else 'Inactivo',
                persona.get('cargo'), persona.get('tipo_contrato'), persona.get('modalidad'),
                persona.get('sueldo'), persona.get('resolucion')
            ]
            ws.append(row_data)

        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

        excel_stream = io.BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        
        return excel_stream

    def check_document_status_for_all_personal(self, days_to_expire=30):
        """Revisa documentos con fecha de vencimiento y resume el estado por persona."""
        all_docs = self._personal_repo.get_all_documents_with_expiration()
        status_summary = {}
        today = datetime.now().date()
        expiration_threshold = today + timedelta(days=days_to_expire)

        for doc in all_docs:
            personal_id = doc['id_personal']
            vencimiento = doc['fecha_vencimiento']

            if personal_id not in status_summary:
                status_summary[personal_id] = {'expired': 0, 'expiring_soon': 0}

            if vencimiento < today:
                status_summary[personal_id]['expired'] += 1
            elif vencimiento <= expiration_threshold:
                status_summary[personal_id]['expiring_soon'] += 1
        
        return status_summary

    def get_expiring_documents_notifications(self, days_threshold=30):
        """
        Orquesta la obtención de una lista de notificaciones sobre documentos que están por vencer.
        """
        return self._personal_repo.find_expiring_documents(days_threshold)

    def get_empleados_por_unidad(self):
        """
        Orquesta la obtención del conteo de empleados por cada unidad administrativa.
        Este método es utilizado por el panel de RRHH para generar gráficos.
        """
        return self._personal_repo.count_empleados_por_unidad()

    def get_empleados_activos_inactivos(self):
        """Orquesta la obtención del conteo de empleados por estado (activo/inactivo)."""
        return self._personal_repo.count_empleados_por_estado()

    def get_empleados_por_sexo(self):
        """Orquesta la obtención del conteo de empleados por sexo."""
        return self._personal_repo.count_empleados_por_sexo()

    def update_personal_details(self, personal_id, form_data, updating_user_id):
        """Actualiza los detalles de un legajo de personal y audita la acción."""
        self._personal_repo.update(personal_id, form_data)
        self._audit_service.log(
            updating_user_id, 
            'Personal', 
            'ACTUALIZAR', 
            f"Se actualizaron los datos del legajo para el personal ID {personal_id}", 
            form_data
        )

    def get_deleted_documents(self):
        """Obtiene todos los documentos marcados como eliminados (activo = 0)."""
        return self._personal_repo.get_deleted_documents()

    def get_document_by_id(self, document_id):
        """Obtiene los detalles de un documento específico."""
        return self._personal_repo.find_document_by_id(document_id)

    def recover_document(self, document_id):
        """Reactiva un documento marcado como eliminado."""
        self._personal_repo.recover_document(document_id)

    def permanently_delete_document(self, document_id):
        """Elimina permanentemente un documento de la base de datos."""
        self._personal_repo.permanently_delete_document(document_id)

    def verify_document_access(self, document_id, user):
        """
        Verifica si un usuario tiene permiso para ver un documento.
        Retorna True si es admin/rrhh o si es el dueño del documento.
        """
        # Roles administrativos tienen acceso total
        if user.rol in ['AdministradorLegajos', 'RRHH', 'Sistemas']:
            return True
            
        # Para personal, verificar propiedad
        if user.rol == 'Personal':
            owner_id = self._personal_repo.get_document_owner(document_id)
            user_personal_id = getattr(user, 'id_personal', None)
            
            # Si el documento existe y pertenece al usuario actual
            return owner_id is not None and owner_id == user_personal_id
            
        return False

